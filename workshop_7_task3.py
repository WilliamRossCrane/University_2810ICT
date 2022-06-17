#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Sep  3 11:39:17 2020
@author: WillCrane
"""
## - Task 3 - Load and explore an Excel Data file. - ##
import openpyxl
# Opens the Excel File
wb = openpyxl.load_workbook('Stocks.xlsx')
# Returns Sheetname
wb.sheetnames
#Specifys Sheet
sheet = wb['IBM']
# Represents the Specific Worksheet
type(sheet)
# Loops through the Sheet
# If Column A Contains 2010 it will print the Row
for row in sheet.iter_rows(min_row=1, max_row=252, max_col=7):
        for cell in row:
            print(cell.value)
    