#-- TASK 3 --#
# Install: pip3 install openpyxl
# Import Load_workbook to Access Excel Files
from openpyxl import load_workbook
# Module to Allow Work with Date and Time
from datetime import datetime
# Module to Alllow Calculation of Mean/Average
from statistics import mean
# Library to Allow Graph Plotting
import matplotlib.pyplot as plt
# Loads the ClimateData Workbook
wb = load_workbook(filename = 'ClimateData.xlsx')
# Makes the Workbook Active
ws = wb.active
# Creates and Empty Dictionary 
all_tmp = {}
# Loop to Enumerate Through the Workbook
for rowno, rowval in enumerate(ws.iter_rows(min_row=1, min_col=1), start=2):
    # Loop to Enumerate Through Each Cell
    for cell in rowval:
        # If Statement to Break up Each Column (Year, City and Temperature)
        if ws.cell(row=rowno,column=1).value:
            year=ws.cell(row=rowno,column=1).value.strftime('%Y')
            at=ws.cell(row=rowno,column=2).value
            city=ws.cell(row=rowno,column=4).value
            # Checks if Year is Between 1900 and 2000
            if int(year) >= 1900 and int(year) <= 2000:
                # If True it Appends the Dictionary to Add the City  
                if all_tmp.get(city):
                    all_tmp[city].append(at)
                # Else No City is Added
                else:
                    all_tmp[city] = []
                    all_tmp[city].append(at)
# Finds the Mean Temp of Each City 
syt = mean(all_tmp['Sydney'])
lot = mean(all_tmp['London'])
nyt = mean(all_tmp['New York'])
jkt = mean(all_tmp['Jakarta'])
# Y Axis Contains the City Temparture Varaibles 
y = [syt,lot,nyt,jkt]
# X Axis Contains the City String Titles
x = ['Sydney','London','New York','Jakarta']
# Creates a New Figure
plt.figure()
# Plots the X and Y Variables
plt.plot(x,y)
# Displays the Graph
plt.show()

